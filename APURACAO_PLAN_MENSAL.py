#LER ARQUIVO SPED
#EXTRAIR INFORMANÇÕES C190
#FAZER LISTA E SOMAR OS VALOR

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

arquivo = r"C:\Users\CLAY\Documents\EMUSA\APURACAO_2023\SPED 122023.txt"
#arquivo = r"C:\Users\CLAY\Downloads\SpedEFD-19953965000127-353165150111-Remessa de arquivo substituto-jun2020.txt"

##### LER ARQUIVO DE SPED #####
with open (arquivo , 'r+' , encoding='ANSI') as arquivo:
    lst_cfop_soma = {}
    for linha in arquivo:
        if linha.startswith('|C190|') or linha.startswith('|D190|') :
            lst_linha = linha.split('|')
            if linha.startswith('|C190|'):
                var_ipi = float(lst_linha[11].replace(',','.'))
            else:
                var_ipi = float(lst_linha[8].replace(',','.'))
            if lst_linha[3] in lst_cfop_soma:
                #print(lst_cfop_soma[lst_linha[3]])
                print(lst_linha)
                #lst_cfop_soma[lst_linha[3]] += float(lst_linha[5].replace(',','.')) #6 , 7 ,11
                lst_cfop_soma[lst_linha[3]][0] += float(lst_linha[5].replace(',','.')) #6 , 7 ,11
                lst_cfop_soma[lst_linha[3]][1] += float(lst_linha[6].replace(',','.')) #6 , 7 ,11
                lst_cfop_soma[lst_linha[3]][2] += float(lst_linha[7].replace(',','.')) #6 , 7 ,11
                lst_cfop_soma[lst_linha[3]][3] += var_ipi #6 , 7 ,11
            else:
                #lst_cfop_soma.update({lst_linha[3]:float(lst_linha[5].replace(',','.'))})
                lst_cfop_soma.update({lst_linha[3]:[float(lst_linha[5].replace(',','.')),float(lst_linha[6].replace(',','.')),float(lst_linha[7].replace(',','.')),var_ipi]})
    print(*lst_cfop_soma.items(), sep='\n')


##### IMPORTANDO LISTA PARA PANDAS  ##########

df_cfop1 = pd.DataFrame.from_dict(lst_cfop_soma, orient='index')
df_cfop1 = df_cfop1.sort_index(axis=0).reset_index()
df_cfop1 = df_cfop1.rename(columns={'index':"CFOP",0:"TOTAL_OP",1:'TOTAL_BC_ICMS',2:'TOTAL_ICMS',3:'TOTAL_IPI'})


#### CARREGANDO A PLANILHA EXCEL VAZIA  ######

wb = load_workbook(r'C:\Users\CLAY\Documents\EMUSA\APURACAO_2023\EMUSA_APURACAO_.xlsx')
wb.active = wb['RESUMO_CFOP']

### FUNÇÃO PREENCHER VALORES NA PLANILHA APURAÇÃO
def add_valores( valor1, valor2, valor3, valor4):
    #print(valor1, valor2, valor3, valor4)
    plan1 = wb['RESUMO_CFOP']
    plan2 = wb['APURACAO_ICMS']
    plan3 = wb['APURACAO_IPI']
    cell_range = plan1['B5:B71'] + plan1['G5:G71']
    for i, x in enumerate(cell_range):
        #print(type(x[0].value), x, i)
        #print(type(valor1))
        if int(valor1) == x[0].value:
            if int(valor1) > 5000:
                plan1[f'I{i-62}'] = valor2
                plan2[f'I{i-62}'] = valor3
                plan3[f'I{i-62}'] = valor4
            else:
                plan1[f'D{i+5}'] = valor2
                plan2[f'D{i+5}'] = valor3
                plan3[f'D{i+5}'] = valor4

##########################################

##### APLICANDO A FUNÇÃO NO PANDAS #####
df_cfop1.apply(lambda x: add_valores(x['CFOP'],x['TOTAL_OP'],x['TOTAL_ICMS'],x['TOTAL_IPI']), axis=1)

##### SALVANDO PLANILHA DE EXCEL PREENCHIHDA  ###########
wb.save('EMUSA_APURACAO_.xlsx')
